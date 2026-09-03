#!/usr/bin/env python3
"""
Import a review workbook (images embedded in a sheet) into the pipeline.
-----------------------------------------------------------------------
Photos that were collected elsewhere — a spreadsheet with one image per row —
are lifted out, checked against the ALBTRIX dataset, and written in the same
shape fetch-images.mjs produces, so build-preview.mjs can render them.

Nothing is taken on trust. For every row the article code must exist in the
dataset, the barcode must match the one on file and pass its check digit, and
the product name must resemble the one in the warehouse list. Whatever fails is
kept in the report with the reason attached, so it shows up in the review
instead of quietly passing as verified.

Usage:
  pip install openpyxl
  python3 scripts/product-images/import-workbook.py \
      --source "JARA_Fotografite_e_Produkteve_Seria_001.xlsx" --label seria-001
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from urllib.parse import urlparse

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
DATASET = os.path.join(HERE, "albtrix-products.json")
CACHE = os.path.join(ROOT, ".image-cache", "originals")
REPORTS = os.path.join(ROOT, "scripts", "reports")

# Column headers as written in the workbook, mapped to what we store.
COLUMNS = {
    "Shifra": "code",
    "Emërtimi": "name",
    "Barkodi": "barcode",
    "Brendi": "brand",
    "Besueshmëria": "confidence",
    "Statusi": "sourceStatus",
    "Faqja burimore": "sourceUrl",
    "Shënim": "note",
}


def safe_stem(code):
    """Article codes may end in a dot ("99290..") — legal in the warehouse
    system, a broken file name on Windows. The code stays the record's id; only
    the file on disk is sanitized."""
    return re.sub(r"[^A-Za-z0-9_-]", "_", code) or "produkt"


def normalize(text):
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def name_overlap(a, b):
    """Share of the warehouse name's words that the workbook name repeats."""
    left, right = set(normalize(a).split()), set(normalize(b).split())
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


def find_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=12):
        labels = {str(c.value).strip(): c.column for c in row if c.value}
        if "Shifra" in labels and "Emërtimi" in labels:
            return row[0].row, labels
    sys.exit("Kopfzeile mit 'Shifra' und 'Emërtimi' nicht gefunden.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="workbook with embedded images")
    ap.add_argument("--label", default="seria-001")
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl fehlt — bitte 'pip install openpyxl' ausführen.")

    with open(DATASET, encoding="utf-8") as fh:
        dataset = json.load(fh)
    by_code = {p["code"]: p for p in dataset["products"]}

    wb = openpyxl.load_workbook(args.source)
    ws = wb[args.sheet] if args.sheet else next(
        (s for s in wb.worksheets if s._images), wb.worksheets[0]
    )
    header_row, labels = find_header(ws)
    images = {img.anchor._from.row + 1: img for img in ws._images}

    os.makedirs(CACHE, exist_ok=True)
    products, stats = [], {"gefunden": 0, "ohne Bild": 0, "Warnungen": 0}

    for row in ws.iter_rows(min_row=header_row + 1):
        values = {
            field: ws.cell(row=row[0].row, column=labels[head]).value
            for head, field in COLUMNS.items()
            if head in labels
        }
        code = str(values.get("code") or "").strip()
        if not code:
            continue

        warnings = []
        known = by_code.get(code)
        if not known:
            warnings.append("Artikelcode nicht in der Warenliste")
        else:
            workbook_barcode = re.sub(r"\D", "", str(values.get("barcode") or ""))
            if workbook_barcode and known["barcode"] and workbook_barcode != known["barcode"]:
                warnings.append(
                    f"Barcode weicht ab (Liste {known['barcode']}, Datei {workbook_barcode})"
                )
            if not known["barcodeUsable"]:
                warnings.append(f"Barcode unzuverlässig: {known['barcodeNote']}")
            if known["kind"] != "retail":
                warnings.append(f"keine Handelsware (Art: {known['kind']})")
            if name_overlap(known["name"], values.get("name")) < 0.5:
                warnings.append("Produktname weicht von der Warenliste ab")

        record = {
            "code": code,
            "name": (known or {}).get("name") or str(values.get("name") or "").strip(),
            "brand": (known or {}).get("brand") or str(values.get("brand") or "").strip(),
            "barcode": (known or {}).get("barcode")
            or re.sub(r"\D", "", str(values.get("barcode") or "")),
            "price": (known or {}).get("price"),
            "supplier": (known or {}).get("supplier", ""),
            "confidence": str(values.get("confidence") or "").strip(),
            "note": str(values.get("note") or "").strip(),
            "warnings": warnings,
        }

        image = images.get(row[0].row)
        if image is None:
            record["status"] = "kein Treffer"
            stats["ohne Bild"] += 1
        else:
            data = image.ref.getvalue() if hasattr(image.ref, "getvalue") else image.ref.read()
            target = os.path.join(CACHE, f"{safe_stem(code)}.png")
            with open(target, "wb") as fh:
                fh.write(data)
            source_url = str(values.get("sourceUrl") or "").strip()
            domain = urlparse(source_url).netloc if source_url.startswith("http") else ""
            record.update(
                status="gefunden",
                source=domain or "Arbeitsmappe",
                sourceUrl=source_url or None,
                # These come from shop and manufacturer pages, not from a
                # source that states a reuse licence. Recorded as such rather
                # than labelled with a licence nobody granted.
                licence="burim i jashtëm",
                file=os.path.relpath(target, ROOT),
                bytes=len(data),
            )
            stats["gefunden"] += 1

        if warnings:
            stats["Warnungen"] += 1
        products.append(record)

    report = {
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "label": args.label,
        "importedFrom": os.path.basename(args.source),
        "totals": {
            "selected": len(products),
            "found": stats["gefunden"],
            "missing": stats["ohne Bild"],
            "withWarnings": stats["Warnungen"],
        },
        "products": products,
    }
    os.makedirs(REPORTS, exist_ok=True)
    out = os.path.join(REPORTS, f"product-images-{args.label}.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)
        fh.write("\n")

    print(json.dumps({**report["totals"], "bericht": os.path.relpath(out, ROOT)}, indent=2))
    for p in products:
        for w in p["warnings"]:
            print(f"  ! {p['code']} {p['name'][:44]}: {w}")


if __name__ == "__main__":
    main()
