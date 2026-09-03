#!/usr/bin/env python3
"""
ALBTRIX warehouse export -> clean product dataset for the image fetcher.
------------------------------------------------------------------------
Reads the pharmacy's ALBTRIX export (article code, name, barcode, price,
supplier, brand) and writes `albtrix-products.json` next to this file: one
normalized record per product, with the barcode actually verified rather
than assumed.

Why this exists as a script and not a one-off: the export is re-generated
whenever the assortment changes, and the repo deliberately carries no
xlsx dependency (see scripts/import-shemo.mjs — zero third-party imports).
So the Excel -> JSON step runs here, once, with openpyxl, and everything
downstream reads only the JSON.

Note: *.xlsx is gitignored repo-wide, so the source export itself is never
committed; pass its path explicitly.

Usage:
  pip install openpyxl
  python3 scripts/product-images/prepare-dataset.py \
      --source "/path/to/Lista e Produkteve_me Brende_JARA.xlsx"
  python3 scripts/product-images/prepare-dataset.py --source ... --stats-only
"""

import argparse
import collections
import json
import os
import re
import sys
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(HERE, "albtrix-products.json")

# GS1 prefixes that are NOT globally unique: 02x/04x and 2xx are reserved for
# restricted distribution — in-store labels a pharmacy prints itself. Their
# check digit is valid, but no external database can resolve them, so they are
# useless for image lookup and must not be graded "easy".
RESTRICTED_PREFIXES = ("02", "04", "2")

# ALBTRIX "Lloji i artikullit" (item type) — the export is a bookkeeping
# extract, not a product list: alongside the retail assortment it carries the
# pharmacy's medicine range and even its expense/asset accounts (rent, company
# cars). Only "100" is the OTC shelf we photograph.
ITEM_KINDS = {
    "100": "retail",   # OTC / retail assortment — the images we want
    "901": "pharma",   # medicine range, incl. prescription-only — never auto-published
    "700": "account",  # expense accounts (rent, insurance, fuel)
    "200": "asset",    # fixed assets (company cars, air conditioning)
    "750": "account",
}


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def digits(value):
    return re.sub(r"\D", "", str(value or ""))


def check_digit_ok(code):
    """GS1 mod-10 check digit, for EAN-8 / UPC-A / EAN-13 / GTIN-14."""
    if len(code) not in (8, 12, 13, 14) or not code.isdigit():
        return False
    body = [int(c) for c in code[:-1]][::-1]
    total = sum(d * (3 if i % 2 == 0 else 1) for i, d in enumerate(body))
    return (10 - total % 10) % 10 == int(code[-1])


def barcode_verdict(raw):
    """Return (normalized, usable, reason). `usable` means: an external
    database could plausibly resolve this code to a real product."""
    code = digits(raw)
    if not code:
        return "", False, "kein Barcode"
    if len(code) not in (8, 12, 13, 14):
        return code, False, f"ungewöhnliche Länge ({len(code)} Stellen)"
    if len(set(code)) <= 2:
        return code, False, "Platzhalter (Ziffernwiederholung)"
    if code in ("12345670", "1234567890128"):
        return code, False, "Platzhalter"
    if not check_digit_ok(code):
        return code, False, "Prüfziffer stimmt nicht"
    normalized = code[1:] if len(code) == 13 and code[0] == "0" else code
    if normalized.startswith(RESTRICTED_PREFIXES):
        return code, False, "hausinterner Code (nicht weltweit eindeutig)"
    return code, True, "gültig"


def build_brand_vocabulary(rows, brand_col):
    """Brands spelled out in the Marke column, longest first, so that
    'NATURES TRUTH' wins over 'NATURES' when scanning a product name."""
    vocab = {clean_text(r[brand_col]).upper() for r in rows if clean_text(r[brand_col])}
    vocab.discard("")
    return sorted(vocab, key=len, reverse=True)


def brand_from_name(name, vocabulary):
    """Recover the brand from the product name. Most rows without a Marke
    entry still start with it: 'ACNECINAMIDE CLEANSING GEL 150 ml'."""
    upper = f" {clean_text(name).upper()} "
    for brand in vocabulary:
        if upper.startswith(f" {brand} "):
            return brand
    return ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="ALBTRIX .xlsx export")
    ap.add_argument("--sheet", default="ALBTRIX Export")
    ap.add_argument("--stats-only", action="store_true")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl fehlt — bitte 'pip install openpyxl' ausführen.")

    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [clean_text(h) for h in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    required = [
        "Shifra",
        "Emërtimi",
        "Barkodi",
        "Lloji i artikullit",
        "Çmimi shitës",
        "Furnitori",
        "Brendi",
    ]
    missing = [c for c in required if c not in col]
    if missing:
        sys.exit(f"Spalten fehlen im Export: {missing} (gefunden: {header})")

    body = [r for r in rows[1:] if any(c is not None for c in r)]
    vocabulary = build_brand_vocabulary(body, col["Brendi"])

    products, seen_codes = [], set()
    for row in body:
        code = clean_text(row[col["Shifra"]])
        name = clean_text(row[col["Emërtimi"]])
        if not name or code in seen_codes:
            continue
        seen_codes.add(code)

        barcode, usable, reason = barcode_verdict(row[col["Barkodi"]])
        brand = clean_text(row[col["Brendi"]]).upper()
        brand_source = "spalte" if brand else ""
        if not brand:
            brand = brand_from_name(name, vocabulary)
            brand_source = "name" if brand else ""

        price = row[col["Çmimi shitës"]]
        item_type = clean_text(row[col["Lloji i artikullit"]])
        products.append(
            {
                "code": code,
                "name": name,
                "itemType": item_type,
                "kind": ITEM_KINDS.get(item_type, "unknown"),
                "barcode": barcode,
                "barcodeUsable": usable,
                "barcodeNote": reason,
                "brand": brand,
                "brandSource": brand_source,
                "price": float(price) if isinstance(price, (int, float)) else None,
                "supplier": clean_text(row[col["Furnitori"]]).strip('" '),
            }
        )

    retail = [p for p in products if p["kind"] == "retail"]
    counts = {
        "rows": len(products),
        "byKind": dict(collections.Counter(p["kind"] for p in products)),
        "retail": len(retail),
        "retailWithBrand": sum(1 for p in retail if p["brand"]),
        "brandRecoveredFromName": sum(1 for p in products if p["brandSource"] == "name"),
        "retailWithUsableBarcode": sum(1 for p in retail if p["barcodeUsable"]),
        "retailReadyForLookup": sum(
            1 for p in retail if p["barcodeUsable"] and p["brand"]
        ),
        "barcodeNotes": dict(collections.Counter(p["barcodeNote"] for p in products)),
    }

    print(json.dumps(counts, indent=2, ensure_ascii=False))
    if args.stats_only:
        return

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": os.path.basename(args.source),
        "counts": counts,
        "products": products,
    }
    with open(OUT_FILE, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
        fh.write("\n")
    print(f"geschrieben: {OUT_FILE}")


if __name__ == "__main__":
    main()
